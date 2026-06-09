#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
과제별 엑셀(data/*.xlsx) → index.html 동기화 도구.

사용법:
  python sync_excel.py --init   # data/ 엑셀 5개를 현재 데이터로 생성 (최초 1회)
  python sync_excel.py          # data/ 엑셀을 읽어 index.html을 갱신

과제 엑셀(총괄·세부1·2·3) 시트:
  - 업무리스트: 구분키 | 구분명 | 업무명 | 계획시작 | 계획종료 | 실적시작 | 실적종료 | 비고 | 출처 | 주관
      · 계획/실적 날짜로 간트에 바 2개(계획 위·실적 아래)가 그려진다. 실적은 비워도 됨.
  - 성능목표실적: 평가항목 | 단위 | 목표 | 실적 | 달성도 | 비고
  - 정량성과:     성과구분 | 단위 | 목표 | 실적 | 비고
  - 간트차트: 자동 생성(편집 금지)
index.html의 GEN:CATS / GEN:SUBPROJECTS / GEN:PERF_CHONG / GEN:QUANT_CHONG /
GEN:CHECKLIST / GEN:SUPPLIES / GEN:TIMELINE 영역이 재생성된다.
"""
import os, re, sys, json, datetime, calendar
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment

BASE = os.path.dirname(os.path.abspath(__file__))
INDEX = os.path.join(BASE, "index.html")
DATA = os.path.join(BASE, "data")

LIST_SHEET = "업무리스트"
PERF_SHEET = "성능목표실적"
QUANT_SHEET = "정량성과"
GANTT_SHEET = "간트차트"
LIST_HEADERS = ["구분키", "구분명", "업무명", "계획시작", "계획종료", "실적시작", "실적종료", "비고", "출처", "주관"]
PERF_HEADERS = ["평가항목", "단위", "목표", "실적", "달성도", "비고"]
QUANT_HEADERS = ["성과구분", "단위", "목표", "실적", "비고"]
CLS = ["a", "b", "c", "d"]
CLS_HEX = {"a": "2563EB", "b": "7A4FD0", "c": "15A06A", "d": "D9822B"}
GANTT_START = (2026, 1)
GANTT_END = (2027, 3)
SRC_TO_KO = {"doc": "계획서", "std": "행정"}
KO_TO_SRC = {"계획서": "doc", "행정": "std", "doc": "doc", "std": "std", "계획": "doc"}
JUK_TO_KO = {"chong": "총괄", "op": "운영위", "dlr": "DLR", "ketep": "전담"}
KO_TO_JUK = {"총괄": "chong", "운영위": "op", "DLR": "dlr", "dlr": "dlr",
             "전담": "ketep", "chong": "chong", "op": "op", "ketep": "ketep"}

# ── 과제별 표시 메타데이터 (제목/아이콘만; 업무 내용은 엑셀에서) ───────────────
PROJECTS = [
    {"file": "총괄.xlsx", "target": "CATS", "id": "chong", "with_juk": True},
    {"file": "세부1.xlsx", "target": "SUB", "id": "sub1", "prefix": "u1", "nav": "🔬 세부1과제",
     "title": "세부1과제 · 2026년(2차년도) 수행 업무",
     "sub": "(편집 필요) 세부1과제 수행 업무와 일정 — 예시를 실제 내용으로 교체하세요",
     "chips": ["🧩 <b>세부1</b> 과제", "🤝 주관 <b>KIMM</b>", "🗓️ 2차년도 <b>2026</b>"]},
    {"file": "세부2.xlsx", "target": "SUB", "id": "sub2", "prefix": "u2", "nav": "🧪 세부2과제",
     "title": "세부2과제 · 2026년(2차년도) 수행 업무",
     "sub": "(편집 필요) 세부2과제 수행 업무와 일정 — 예시를 실제 내용으로 교체하세요",
     "chips": ["🧩 <b>세부2</b> 과제", "🤝 주관 <b>KIMM</b>", "🗓️ 2차년도 <b>2026</b>"]},
    {"file": "세부3.xlsx", "target": "SUB", "id": "sub3", "prefix": "u3", "nav": "⚙️ 세부3과제",
     "title": "세부3과제 · 2026년(2차년도) 수행 업무",
     "sub": "(편집 필요) 세부3과제 수행 업무와 일정 — 예시를 실제 내용으로 교체하세요",
     "chips": ["🧩 <b>세부3</b> 과제", "🤝 주관 <b>KIMM</b>", "🗓️ 2차년도 <b>2026</b>"]},
]

# ── 최초 시드 데이터 (--init 으로 엑셀 생성 시 사용) ──────────────────────────
SEED = {
    "chong": [
        {"key": "A", "title": "사업 운영·관리", "tasks": [
            {"name": "2차년도 협약 체결·연구비 카드 등록·집행계획 수립", "juk": ["chong"], "start": "2026-01-02", "end": "2026-03-12", "astart": "2026-01-05", "aend": "2026-03-20", "note": "협약 3/12 기준 (실적 예시)", "src": "std"},
            {"name": "참여연구원 등록·변경 및 인건비 계상 관리", "juk": ["chong"], "start": "2026-01-05", "end": "2026-12-31", "note": "연중 상시", "src": "std"},
            {"name": "운영위원회 개최 (연 1회 이상)", "juk": ["op"], "start": "2026-09-01", "end": "2026-09-30", "note": "KETEP 간사·국내전문가(가스터빈·수전해) 포함", "src": "doc"},
            {"name": "세부과제(1~5) 진척도 정기 점검 회의", "juk": ["chong"], "start": "2026-03-01", "end": "2026-12-31", "note": "분기별 (3·6·9·12월)", "src": "doc"},
            {"name": "KPI 기반 진척도 모니터링·조기경보(Early Warning)", "juk": ["chong"], "start": "2026-01-05", "end": "2026-12-31", "note": "연차·분기 보고 체계", "src": "doc"},
            {"name": "외부 전문가 Peer-review (내부 모니터링 병행)", "juk": ["chong"], "start": "2026-10-01", "end": "2026-11-30", "note": "", "src": "doc"},
            {"name": "부진과제 컨설팅·개선계획 수립 (필요시)", "juk": ["chong"], "start": "2026-11-01", "end": "2026-12-15", "note": "조건부 수행", "src": "doc"},
            {"name": "연구노트·보안등급·연구윤리 관리", "juk": ["chong"], "start": "2026-01-05", "end": "2026-12-31", "note": "보안등급 분류 기준", "src": "doc"},
        ]},
        {"key": "B", "title": "국제협력 (KIMM–DLR)", "tasks": [
            {"name": "KIMM–DLR 협력채널 유지·실무 협의", "juk": ["chong", "dlr"], "start": "2026-01-05", "end": "2026-12-31", "note": "내부협의체 정기 운영", "src": "doc"},
            {"name": "국제 워크숍 — 한국(기계연) 준비·개최", "juk": ["chong"], "start": "2026-06-16", "end": "2026-08-05", "note": "8/5 개최 (워크숍 탭 참조)", "src": "doc"},
            {"name": "국제 워크숍/기술교류 — 독일(DLR) 방문", "juk": ["chong", "dlr"], "start": "2026-10-01", "end": "2026-11-15", "note": "한·독 각 1회 이상", "src": "doc"},
            {"name": "연구인력 교류 — DLR 파견·방문연구·세미나 지원", "juk": ["chong", "dlr"], "start": "2026-04-01", "end": "2026-12-31", "note": "하반기 집중", "src": "doc"},
            {"name": "EKC 한-유럽 과학기술학술회의 특별세션 개설·발표", "juk": ["chong"], "start": "2026-05-01", "end": "2026-07-31", "note": "에너지/기계 분과 세션", "src": "doc"},
        ]},
        {"key": "C", "title": "성과 관리", "tasks": [
            {"name": "공동 논문(공동저자)·특허(공동출원) 성과 관리", "juk": ["chong", "dlr"], "start": "2026-03-01", "end": "2026-12-31", "note": "공동소유·활용 계약 기반", "src": "doc"},
            {"name": "DLR 공동특허 출원·기술이전 체계 운영", "juk": ["chong", "dlr"], "start": "2026-06-01", "end": "2026-12-31", "note": "", "src": "doc"},
            {"name": "수요기업 사업화 설명회 (단계별 1회)", "juk": ["chong"], "start": "2026-09-01", "end": "2026-10-31", "note": "국내 수요기업 대상", "src": "doc"},
            {"name": "성과집·기술 브로슈어 발간", "juk": ["chong"], "start": "2026-11-01", "end": "2026-12-31", "note": "", "src": "doc"},
        ]},
        {"key": "D", "title": "평가·보고 (연차 마감)", "tasks": [
            {"name": "2차년도 연차실적보고서 작성·제출", "juk": ["chong"], "start": "2026-12-01", "end": "2027-02-15", "note": "KETEP 제출", "src": "std"},
            {"name": "2차년도 연구비 집행 점검·정산", "juk": ["chong"], "start": "2027-01-02", "end": "2027-02-28", "note": "", "src": "std"},
            {"name": "자체평가·연차평가(점검) 대응", "juk": ["chong"], "start": "2026-12-15", "end": "2027-01-31", "note": "", "src": "std"},
            {"name": "3차년도 사업계획서·협약 준비", "juk": ["chong"], "start": "2026-11-01", "end": "2027-03-15", "note": "차년도 연속 수행", "src": "std"},
        ]},
    ],
}
for n in (1, 2, 3):
    SEED[f"sub{n}"] = [
        {"key": "1", "title": "(예시) 업무 구분 1", "tasks": [
            {"name": f"(예시) 세부{n} 핵심 연구·실험 항목", "juk": [], "start": "2026-01-05", "end": "2026-06-30", "astart": "2026-01-12", "aend": "2026-05-20", "note": "예시 — 실제 업무로 교체 (실적바 예시)", "src": "doc"},
            {"name": f"(예시) 세부{n} 중간 점검·보고", "juk": [], "start": "2026-07-01", "end": "2026-09-30", "note": "예시", "src": "std"},
        ]},
    ]

# 성능 목표·실적 / 정량 성과 시드 (예시) — 과제별 동일 예시, 실제 내용으로 교체
SEED_PERF = {pid: [
    {"item": "(예시) 평가항목 1", "unit": "—", "target": "(목표값)", "actual": "(실적값)", "rate": "", "note": "예시 — 실제 내용으로 교체"},
    {"item": "(예시) 평가항목 2", "unit": "%", "target": "90", "actual": "", "rate": "", "note": "예시"},
] for pid in ("chong", "sub1", "sub2", "sub3")}
SEED_QUANT = {pid: [
    {"name": "SCI(E) 논문", "unit": "편", "target": "1", "actual": "0", "note": "예시"},
    {"name": "특허 출원(국내·국제)", "unit": "건", "target": "1", "actual": "0", "note": "예시"},
    {"name": "학술회의 발표", "unit": "건", "target": "2", "actual": "0", "note": "예시"},
] for pid in ("chong", "sub1", "sub2", "sub3")}


# ───────────────────────── helpers ─────────────────────────
def js(s):
    return json.dumps(s, ensure_ascii=False)


def fmt_date(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    m = re.match(r"^(\d{4})[-./](\d{1,2})[-./](\d{1,2})", s)
    if m:
        return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    return s


def is_date(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return True
    if v is None:
        return False
    return bool(re.match(r"^\d{4}[-./]\d{1,2}[-./]\d{1,2}", str(v).strip()))


def norm_src(v):
    if v is None:
        return None
    return KO_TO_SRC.get(str(v).strip())


def norm_juk(v):
    if not v:
        return []
    out = []
    for tok in str(v).replace("，", ",").split(","):
        tok = tok.strip()
        if tok and tok in KO_TO_JUK and KO_TO_JUK[tok] not in out:
            out.append(KO_TO_JUK[tok])
    return out


def _s(v):
    return "" if v is None else str(v).strip()


# ───────────────────────── read xlsx ─────────────────────────
def read_ws(ws):
    cats, by_key = [], {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None:
            continue
        vals = (list(row) + [None] * 10)[:10]
        key, name, task, pstart, pend, astart, aend, note, src, juk = vals
        key = _s(key)
        task = _s(task)
        if not task:
            continue
        if key not in by_key:
            cat = {"key": key or str(len(cats) + 1), "cls": CLS[len(cats) % 4],
                   "title": (_s(name) or key), "tasks": []}
            by_key[key] = cat
            cats.append(cat)
        t = {"name": task, "start": fmt_date(pstart), "end": fmt_date(pend)}
        if is_date(astart) and is_date(aend):
            t["astart"] = fmt_date(astart)
            t["aend"] = fmt_date(aend)
        if _s(note):
            t["note"] = _s(note)
        s = norm_src(src)
        if s:
            t["src"] = s
        j = norm_juk(juk)
        if j:
            t["juk"] = j
        by_key[key]["tasks"].append(t)
    return cats


def read_perf(ws):
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        item, unit, target, actual, rate, note = (list(row) + [None] * 6)[:6]
        if not _s(item):
            continue
        out.append({"item": _s(item), "unit": _s(unit), "target": _s(target),
                    "actual": _s(actual), "rate": _s(rate), "note": _s(note)})
    return out


def read_quant(ws):
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name, unit, target, actual, note = (list(row) + [None] * 5)[:5]
        if not _s(name):
            continue
        out.append({"name": _s(name), "unit": _s(unit), "target": _s(target),
                    "actual": _s(actual), "note": _s(note)})
    return out


# ───────────────────────── data → JS ─────────────────────────
def emit_task(t, with_juk):
    parts = [f"name:{js(t['name'])}"]
    if with_juk and t.get("juk"):
        parts.append("juk:[" + ",".join(js(x) for x in t["juk"]) + "]")
    parts += [f"start:{js(t['start'])}", f"end:{js(t['end'])}"]
    if t.get("astart") and t.get("aend"):
        parts += [f"astart:{js(t['astart'])}", f"aend:{js(t['aend'])}"]
    if t.get("note"):
        parts.append(f"note:{js(t['note'])}")
    if t.get("src"):
        parts.append(f"src:{js(t['src'])}")
    return "{ " + ", ".join(parts) + " }"


def emit_cats(cats, with_juk, pad):
    blocks = []
    for cat in cats:
        tl = (",\n").join(pad + "    " + emit_task(t, with_juk) for t in cat["tasks"])
        blocks.append(f"{pad}{{ key:{js(cat['key'])}, cls:{js(cat['cls'])}, "
                      f"title:{js(cat['title'])}, tasks:[\n{tl} ]}}")
    return "[\n" + ",\n".join(blocks) + " ]"


def emit_perf(perf):
    rows = ["    { item:%s, unit:%s, target:%s, actual:%s, rate:%s, note:%s }" %
            (js(p["item"]), js(p["unit"]), js(p["target"]), js(p["actual"]), js(p["rate"]), js(p["note"]))
            for p in perf]
    return "[\n" + ",\n".join(rows) + " ]" if rows else "[]"


def emit_quant(quant):
    rows = ["    { name:%s, unit:%s, target:%s, actual:%s, note:%s }" %
            (js(q["name"]), js(q["unit"]), js(q["target"]), js(q["actual"]), js(q["note"]))
            for q in quant]
    return "[\n" + ",\n".join(rows) + " ]" if rows else "[]"


def build_cats_js(cats):
    return "const CATS = " + emit_cats(cats, True, "  ") + ";"


def build_perf_chong_js(perf):
    return "const PERF_CHONG = " + emit_perf(perf) + ";"


def build_quant_chong_js(quant):
    return "const QUANT_CHONG = " + emit_quant(quant) + ";"


def build_subprojects_js(sub_defs):
    objs = []
    for p, cats, perf, quant in sub_defs:
        chips = "[" + ",".join(js(c) for c in p["chips"]) + "]"
        objs.append(
            f"  {{ id:{js(p['id'])}, prefix:{js(p['prefix'])}, nav:{js(p['nav'])},\n"
            f"    title:{js(p['title'])},\n"
            f"    sub:{js(p['sub'])},\n"
            f"    chips:{chips},\n"
            f"    cats:{emit_cats(cats, False, '    ')},\n"
            f"    perf:{emit_perf(perf)},\n"
            f"    quant:{emit_quant(quant)} }}")
    return "const SUBPROJECTS = [\n" + ",\n".join(objs) + "\n];"


def inject(html, marker, new_code):
    pat = re.compile(r"(/\* GEN:%s:START \*/\n).*?(\n/\* GEN:%s:END \*/)" % (marker, marker), re.DOTALL)
    if not pat.search(html):
        sys.exit(f"[오류] index.html에서 GEN:{marker} 마커를 찾을 수 없습니다.")
    return pat.sub(lambda m: m.group(1) + new_code + m.group(2), html)


# ───────────────────────── gantt 미리보기 시트 (자동 생성) ─────────────────────────
def month_list():
    out, (y, m), (ey, em) = [], GANTT_START, GANTT_END
    while (y, m) <= (ey, em):
        out.append((y, m))
        m += 1
        if m > 12:
            y, m = y + 1, 1
    return out


def _date(s):
    y, m, d = (int(x) for x in s.split("-"))
    return datetime.date(y, m, d)


def write_gantt_sheet(wb, cats):
    """간트 미리보기(계획 기준). 실적은 HTML 대시보드에서 계획·실적 2바로 표시."""
    if GANTT_SHEET in wb.sheetnames:
        del wb[GANTT_SHEET]
    ws = wb.create_sheet(GANTT_SHEET)
    months = month_list()
    ncols = 3 + len(months)
    head_font = Font(name="Arial", bold=True, color="FFFFFF", size=9)
    body_font = Font(name="Arial", size=9)
    navy = PatternFill("solid", fgColor="16335E")
    thin = Side(style="thin", color="E6EAF0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    ws.cell(1, 1, "⚠ 자동 생성 탭(계획 기준 미리보기) — 직접 편집 금지. 일정은 '업무리스트'에서 수정 후 python sync_excel.py.")
    ws.cell(1, 1).font = Font(name="Arial", size=9, italic=True, color="B00000")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    hdr = ["구분", "업무명", "계획기간"] + [(f"'{str(y)[2:]} {m}월" if m == 1 else f"{m}월") for (y, m) in months]
    for c, h in enumerate(hdr, 1):
        cell = ws.cell(2, c, h)
        cell.fill, cell.font, cell.alignment, cell.border = navy, head_font, center, border
    r = 3
    for ci, cat in enumerate(cats):
        cls = cat.get("cls") or CLS[ci % 4]
        fill = PatternFill("solid", fgColor=CLS_HEX.get(cls, "2563EB"))
        for ti, t in enumerate(cat["tasks"]):
            ws.cell(r, 1, cat["title"] if ti == 0 else "")
            ws.cell(r, 2, t["name"])
            ws.cell(r, 3, f"{t['start']}~{t['end']}")
            s, e = _date(t["start"]), _date(t["end"])
            for mi, (y, m) in enumerate(months):
                cell = ws.cell(r, 4 + mi)
                cell.border = border
                mstart = datetime.date(y, m, 1)
                mend = datetime.date(y, m, calendar.monthrange(y, m)[1])
                if s <= mend and e >= mstart:
                    cell.fill = fill
            for col in (1, 2, 3):
                ws.cell(r, col).font = body_font
                ws.cell(r, col).border = border
            ws.cell(r, 2).alignment = Alignment(vertical="center", wrap_text=True)
            ws.cell(r, 3).alignment = center
            r += 1
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 22
    for i in range(len(months)):
        ws.column_dimensions[ws.cell(2, 4 + i).column_letter].width = 6
    ws.freeze_panes = "D3"
    return ws


# ───────────────────────── write project xlsx (--init) ─────────────────────────
def _hdr_row(ws, headers, comments=None):
    navy = PatternFill("solid", fgColor="16335E")
    hf = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    ce = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="D0D7E2")
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill, cell.font, cell.alignment, cell.border = navy, hf, ce, bd
    for col, txt in (comments or {}).items():
        ws.cell(1, col).comment = Comment(txt, "sync_excel")
    ws.freeze_panes = "A2"
    return bd


def write_xlsx(path, cats, with_juk, perf, quant):
    wb = Workbook()
    body = Font(name="Arial", size=10)
    wrap = Alignment(vertical="center", wrap_text=True)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(vertical="center")

    # 업무리스트 (계획·실적 날짜)
    ws = wb.active
    ws.title = LIST_SHEET
    bd = _hdr_row(ws, LIST_HEADERS, {
        1: "같은 키끼리 한 카테고리로 묶임. 막대 색은 키 등장 순서대로 자동(파랑·보라·초록·주황).",
        4: "계획 시작/종료 — 간트 위쪽(계획) 바.",
        6: "실적 시작/종료 — 간트 아래쪽(실적) 바. 비우면 실적 바 없음.",
        9: "계획서 또는 행정 (비우면 배지 없음).",
        10: "총괄/운영위/DLR/전담 중 쉼표로 구분. 세부과제는 보통 비움.",
    })
    r = 2
    for cat in cats:
        for t in cat["tasks"]:
            ws.cell(r, 1, cat["key"]); ws.cell(r, 2, cat["title"]); ws.cell(r, 3, t["name"])
            for col, kk in ((4, "start"), (5, "end"), (6, "astart"), (7, "aend")):
                if t.get(kk):
                    y, m, d = (int(x) for x in t[kk].split("-"))
                    dc = ws.cell(r, col, datetime.date(y, m, d)); dc.number_format = "yyyy-mm-dd"
            ws.cell(r, 8, t.get("note", ""))
            ws.cell(r, 9, SRC_TO_KO.get(t.get("src"), ""))
            ws.cell(r, 10, ", ".join(JUK_TO_KO[j] for j in t.get("juk", [])) if with_juk else "")
            for col in range(1, 11):
                cc = ws.cell(r, col); cc.font = body; cc.border = bd
                cc.alignment = wrap if col in (3, 8) else center if col in (1, 4, 5, 6, 7, 9) else left
            r += 1
    for i, w in enumerate([8, 18, 46, 12, 12, 12, 12, 24, 9, 14], 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w
    dv = DataValidation(type="list", formula1='"계획서,행정"', allow_blank=True)
    ws.add_data_validation(dv); dv.add(f"I2:I{max(r, 200)}")

    # 성능목표실적
    ws2 = wb.create_sheet(PERF_SHEET)
    bd2 = _hdr_row(ws2, PERF_HEADERS, {5: "예: 85% (수동 입력)."})
    r = 2
    for p in perf:
        for col, k in ((1, "item"), (2, "unit"), (3, "target"), (4, "actual"), (5, "rate"), (6, "note")):
            ws2.cell(r, col, p.get(k, ""))
        for col in range(1, 7):
            cc = ws2.cell(r, col); cc.font = body; cc.border = bd2
            cc.alignment = wrap if col in (1, 6) else center
        r += 1
    for i, w in enumerate([40, 8, 14, 14, 10, 24], 1):
        ws2.column_dimensions[ws2.cell(1, i).column_letter].width = w

    # 정량성과
    ws3 = wb.create_sheet(QUANT_SHEET)
    bd3 = _hdr_row(ws3, QUANT_HEADERS)
    r = 2
    for q in quant:
        for col, k in ((1, "name"), (2, "unit"), (3, "target"), (4, "actual"), (5, "note")):
            ws3.cell(r, col, q.get(k, ""))
        for col in range(1, 6):
            cc = ws3.cell(r, col); cc.font = body; cc.border = bd3
            cc.alignment = wrap if col in (1, 5) else center
        r += 1
    for i, w in enumerate([30, 8, 12, 12, 28], 1):
        ws3.column_dimensions[ws3.cell(1, i).column_letter].width = w

    write_gantt_sheet(wb, cats)
    wb.active = wb[LIST_SHEET]
    wb.save(path)


# ============================ WORKSHOP (워크숍 8/5) ============================
WS_FILE = "워크숍.xlsx"
WS_LIST, WS_SUP, WS_TL = "준비업무", "물품준비", "당일타임라인"
TL_CATS = ["티타임", "오프닝", "발표", "중식", "랩투어", "마무리"]
OWNER_HEX = {"DW": "2563EB", "MK": "15A06A", "DK": "B3741A", "SH": "7A4FD0"}
WS_G_START = datetime.date(2026, 6, 15)
WS_G_END = datetime.date(2026, 8, 16)


def md_str(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return f"{v.month}/{v.day}"
    s = str(v).strip()
    m = re.match(r"^(\d{1,2})\s*[/.\-]\s*(\d{1,2})", s)
    return f"{int(m.group(1))}/{int(m.group(2))}" if m else s


def md_date(s):
    m, d = (int(x) for x in str(s).split("/")[:2])
    return datetime.date(2026, m, d)


def _dash(v):
    s = "" if v is None else str(v).strip()
    return s if s else "—"


def read_checklist(ws):
    cats, by = [], {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        num, title, sub, task, owner, start, due, note, flag = (list(row) + [None] * 9)[:9]
        task = "" if task is None else str(task).strip()
        if not task:
            continue
        key = str(num).strip() if num not in (None, "") else (str(title).strip() if title else "")
        if key not in by:
            try:
                nn = int(num)
            except (TypeError, ValueError):
                nn = len(cats) + 1
            cat = {"num": nn, "title": str(title).strip() if title else key,
                   "sub": str(sub).strip() if sub else "", "items": []}
            by[key] = cat
            cats.append(cat)
        it = {"task": task, "owner": str(owner).strip() if owner else "",
              "start": md_str(start), "due": md_str(due)}
        if note and str(note).strip():
            it["note"] = str(note).strip()
        if flag and str(flag).strip():
            it["flag"] = True
        by[key]["items"].append(it)
    return cats


def read_supplies(ws):
    grps, by = [], {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        grp, name, make, recv, appr, loc = (list(row) + [None] * 6)[:6]
        name = "" if name is None else str(name).strip()
        if not name:
            continue
        gk = str(grp).strip() if grp else (grps[-1]["grp"] if grps else "")
        if gk not in by:
            g = {"grp": gk, "rows": []}
            by[gk] = g
            grps.append(g)
        by[gk]["rows"].append({"name": name, "make": _dash(make), "recv": _dash(recv),
                               "appr": _dash(appr), "loc": str(loc).strip() if loc else ""})
    return grps


def read_timeline(ws):
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        time, cat, content, note, warn = (list(row) + [None] * 5)[:5]
        content = "" if content is None else str(content).strip()
        if not content:
            continue
        t = {"time": str(time).strip() if time else "", "cat": str(cat).strip() if cat else "",
             "content": content}
        if note and str(note).strip():
            t["note"] = str(note).strip()
        if warn and str(warn).strip():
            t["warn"] = True
        out.append(t)
    return out


def build_checklist_js(cats):
    blocks = []
    for c in cats:
        items = []
        for it in c["items"]:
            p = [f"task:{js(it['task'])}", f"owner:{js(it['owner'])}",
                 f"start:{js(it['start'])}", f"due:{js(it['due'])}", f"note:{js(it.get('note',''))}"]
            if it.get("flag"):
                p.append("flag:true")
            items.append("    { " + ", ".join(p) + " }")
        blocks.append(f"  {{ num:{int(c['num'])}, title:{js(c['title'])}, sub:{js(c.get('sub',''))}, items:[\n"
                      + ",\n".join(items) + " ]}")
    return "const CHECKLIST = [\n" + ",\n".join(blocks) + "\n];"


def build_supplies_js(grps):
    blocks = []
    for g in grps:
        rows = ["    { name:%s, make:%s, recv:%s, appr:%s, loc:%s }" %
                (js(r["name"]), js(r["make"]), js(r["recv"]), js(r["appr"]), js(r["loc"])) for r in g["rows"]]
        blocks.append(f"  {{ grp:{js(g['grp'])}, rows:[\n" + ",\n".join(rows) + " ]}")
    return "const SUPPLIES = [\n" + ",\n".join(blocks) + "\n];"


def build_timeline_js(rows):
    out = []
    for t in rows:
        p = [f"time:{js(t['time'])}", f"cat:{js(t['cat'])}", f"content:{js(t['content'])}",
             f"note:{js(t.get('note',''))}"]
        if t.get("warn"):
            p.append("warn:true")
        out.append("  { " + ", ".join(p) + " }")
    return "const TIMELINE = [\n" + ",\n".join(out) + "\n];"


def _ws_head(ws, headers, widths):
    navy = PatternFill("solid", fgColor="16335E")
    hf = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    ce = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="D0D7E2")
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill, cell.font, cell.alignment, cell.border = navy, hf, ce, bd
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w
    ws.freeze_panes = "A2"
    return bd


def write_ws_gantt_sheet(wb, checklist):
    if GANTT_SHEET in wb.sheetnames:
        del wb[GANTT_SHEET]
    ws = wb.create_sheet(GANTT_SHEET)
    weeks, d = [], WS_G_START
    while d <= WS_G_END:
        weeks.append(d)
        d = d + datetime.timedelta(days=7)
    body = Font(name="Arial", size=9)
    thin = Side(style="thin", color="E6EAF0")
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    ncols = 4 + len(weeks)
    ws.cell(1, 1, "⚠ 자동 생성 탭 — 직접 편집하지 마세요. 일정은 '준비업무' 탭에서 수정 후 python sync_excel.py 실행.")
    ws.cell(1, 1).font = Font(name="Arial", size=9, italic=True, color="B00000")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    navy = PatternFill("solid", fgColor="16335E")
    hf = Font(name="Arial", bold=True, color="FFFFFF", size=9)
    ce = Alignment(horizontal="center", vertical="center")
    hdr = ["구분", "업무", "담당", "기간"] + [f"{w.month}/{w.day}" for w in weeks]
    for c, h in enumerate(hdr, 1):
        cell = ws.cell(2, c, h)
        cell.fill, cell.font, cell.alignment, cell.border = navy, hf, ce, bd
    widths = [16, 40, 8, 12] + [5] * len(weeks)
    for i, wd in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(2, i).column_letter].width = wd
    r = 3
    for c in checklist:
        for ti, it in enumerate(c["items"]):
            o = it["owner"].split(",")[0].strip() or "DW"
            fill = PatternFill("solid", fgColor=OWNER_HEX.get(o, "2563EB"))
            ws.cell(r, 1, c["title"] if ti == 0 else "")
            ws.cell(r, 2, it["task"])
            ws.cell(r, 3, it["owner"])
            ws.cell(r, 4, f"{it['start']}~{it['due']}")
            s, e = md_date(it["start"]), md_date(it["due"])
            for wi, w in enumerate(weeks):
                cell = ws.cell(r, 5 + wi)
                cell.border = bd
                if s <= w + datetime.timedelta(days=6) and e >= w:
                    cell.fill = fill
            for col in (1, 2, 3, 4):
                ws.cell(r, col).font = body
                ws.cell(r, col).border = bd
            ws.cell(r, 2).alignment = Alignment(vertical="center", wrap_text=True)
            r += 1
    ws.freeze_panes = "E3"
    return ws


def write_workshop_xlsx(path, checklist, supplies, timeline):
    wb = Workbook()
    body = Font(name="Arial", size=10)
    wrap = Alignment(vertical="center", wrap_text=True)
    ce = Alignment(horizontal="center", vertical="center")
    left = Alignment(vertical="center")
    ws = wb.active
    ws.title = WS_LIST
    bd = _ws_head(ws, ["구분번호", "구분명", "세부설명", "업무", "담당", "시작", "마감", "비고", "⚠표시"],
                  [9, 16, 18, 50, 10, 8, 8, 26, 8])
    r = 2
    for c in checklist:
        for it in c["items"]:
            ws.cell(r, 1, c["num"]); ws.cell(r, 2, c["title"]); ws.cell(r, 3, c.get("sub", ""))
            ws.cell(r, 4, it["task"]); ws.cell(r, 5, it["owner"])
            ws.cell(r, 6, it["start"]); ws.cell(r, 7, it["due"]); ws.cell(r, 8, it.get("note", ""))
            ws.cell(r, 9, "⚠" if it.get("flag") else "")
            for col in range(1, 10):
                cc = ws.cell(r, col); cc.font = body; cc.border = bd
                cc.alignment = wrap if col in (4, 8) else ce if col in (1, 5, 6, 7, 9) else left
            r += 1
    ws2 = wb.create_sheet(WS_SUP)
    bd2 = _ws_head(ws2, ["구분", "물품명", "제작/구매", "수령", "결재", "세팅위치"], [16, 40, 12, 8, 8, 22])
    r = 2
    for g in supplies:
        for row in g["rows"]:
            ws2.cell(r, 1, g["grp"]); ws2.cell(r, 2, row["name"]); ws2.cell(r, 3, row["make"])
            ws2.cell(r, 4, row["recv"]); ws2.cell(r, 5, row["appr"]); ws2.cell(r, 6, row["loc"])
            for col in range(1, 7):
                cc = ws2.cell(r, col); cc.font = body; cc.border = bd2
                cc.alignment = wrap if col == 2 else ce if col in (3, 4, 5) else left
            r += 1
    ws3 = wb.create_sheet(WS_TL)
    bd3 = _ws_head(ws3, ["시간", "구분", "내용", "비고", "⚠확정필요"], [14, 10, 40, 24, 10])
    r = 2
    for t in timeline:
        ws3.cell(r, 1, t["time"]); ws3.cell(r, 2, t["cat"]); ws3.cell(r, 3, t["content"])
        ws3.cell(r, 4, t.get("note", "")); ws3.cell(r, 5, "⚠" if t.get("warn") else "")
        for col in range(1, 6):
            cc = ws3.cell(r, col); cc.font = body; cc.border = bd3
            cc.alignment = wrap if col in (3, 4) else ce if col in (1, 2, 5) else left
        r += 1
    dv = DataValidation(type="list", formula1='"%s"' % ",".join(TL_CATS), allow_blank=True)
    ws3.add_data_validation(dv); dv.add(f"B2:B{max(r, 100)}")
    write_ws_gantt_sheet(wb, checklist)
    wb.active = wb[WS_LIST]
    wb.save(path)


def extract_html_arrays(names):
    node = r"""
const fs=require('fs');
const html=fs.readFileSync(process.argv[1],'utf8');
const s=html.match(/<script>([\s\S]*)<\/script>/)[1];
function arr(src,name){let k=src.indexOf('const '+name);let i=src.indexOf('[',k),depth=0,inStr=null;
 for(let j=i;j<src.length;j++){const c=src[j];
  if(inStr){if(c===inStr&&src[j-1]!=='\\')inStr=null;continue;}
  if(c==='"'||c==="'"||c==='`'){inStr=c;continue;}
  if(c==='[')depth++;else if(c===']'){depth--;if(depth===0)return src.slice(i,j+1);}}}
const out={};process.argv.slice(2).forEach(n=>{out[n]=eval('('+arr(s,n)+')');});
console.log(JSON.stringify(out));
"""
    import subprocess
    res = subprocess.run(["node", "-e", node, INDEX] + names, capture_output=True, text=True, encoding="utf-8")
    if res.returncode != 0:
        sys.exit("[오류] node 데이터 추출 실패:\n" + (res.stderr or res.stdout))
    return json.loads(res.stdout)


def do_init():
    os.makedirs(DATA, exist_ok=True)
    for p in PROJECTS:
        pid = p["id"]
        write_xlsx(os.path.join(DATA, p["file"]), SEED[pid], p.get("with_juk", False),
                   SEED_PERF[pid], SEED_QUANT[pid])
        print(f"  생성: data/{p['file']}  (업무리스트·성능목표실적·정량성과 + 간트)")
    d = extract_html_arrays(["CHECKLIST", "SUPPLIES", "TIMELINE"])
    write_workshop_xlsx(os.path.join(DATA, WS_FILE), d["CHECKLIST"], d["SUPPLIES"], d["TIMELINE"])
    print(f"  생성: data/{WS_FILE}  (준비업무·물품준비·당일타임라인 + 간트)")
    print("초기 엑셀 생성 완료. 이제 내용을 편집한 뒤 `python sync_excel.py`로 반영하세요.")


def _read_project(path):
    wb = load_workbook(path)
    ws = wb[LIST_SHEET] if LIST_SHEET in wb.sheetnames else wb.active
    cats = read_ws(ws)
    perf = read_perf(wb[PERF_SHEET]) if PERF_SHEET in wb.sheetnames else []
    quant = read_quant(wb[QUANT_SHEET]) if QUANT_SHEET in wb.sheetnames else []
    write_gantt_sheet(wb, cats)
    wb.active = wb[LIST_SHEET] if LIST_SHEET in wb.sheetnames else wb.active
    wb.save(path)
    return cats, perf, quant


def do_sync():
    if not os.path.exists(INDEX):
        sys.exit("[오류] index.html 없음")
    html = open(INDEX, encoding="utf-8").read()
    cats_chong = perf_chong = quant_chong = None
    sub_defs, total = [], 0
    for p in PROJECTS:
        path = os.path.join(DATA, p["file"])
        if not os.path.exists(path):
            sys.exit(f"[오류] {path} 없음. 먼저 `python sync_excel.py --init` 실행.")
        cats, perf, quant = _read_project(path)
        n = sum(len(c["tasks"]) for c in cats)
        total += n
        print(f"  읽음: data/{p['file']}  업무 {n}건 · 성능 {len(perf)} · 정량 {len(quant)}")
        if p["target"] == "CATS":
            cats_chong, perf_chong, quant_chong = cats, perf, quant
        else:
            sub_defs.append((p, cats, perf, quant))
    html = inject(html, "CATS", build_cats_js(cats_chong))
    html = inject(html, "PERF_CHONG", build_perf_chong_js(perf_chong))
    html = inject(html, "QUANT_CHONG", build_quant_chong_js(quant_chong))
    html = inject(html, "SUBPROJECTS", build_subprojects_js(sub_defs))
    ws_path = os.path.join(DATA, WS_FILE)
    if os.path.exists(ws_path):
        wb = load_workbook(ws_path)
        checklist = read_checklist(wb[WS_LIST] if WS_LIST in wb.sheetnames else wb.worksheets[0])
        supplies = read_supplies(wb[WS_SUP]) if WS_SUP in wb.sheetnames else []
        timeline = read_timeline(wb[WS_TL]) if WS_TL in wb.sheetnames else []
        write_ws_gantt_sheet(wb, checklist)
        wb.active = wb[WS_LIST] if WS_LIST in wb.sheetnames else wb.active
        wb.save(ws_path)
        html = inject(html, "CHECKLIST", build_checklist_js(checklist))
        html = inject(html, "SUPPLIES", build_supplies_js(supplies))
        html = inject(html, "TIMELINE", build_timeline_js(timeline))
        wn = sum(len(c["items"]) for c in checklist)
        print(f"  읽음: data/{WS_FILE}  준비업무 {wn}건 · 물품 {sum(len(g['rows']) for g in supplies)}건 · 타임라인 {len(timeline)}건")
    open(INDEX, "w", encoding="utf-8").write(html)
    print(f"index.html 갱신 완료 (과제 업무 {total}건 + 성능·정량 + 워크숍). git push(gtfc156) 하면 사이트에 반영됩니다.")


if __name__ == "__main__":
    if "--init" in sys.argv:
        do_init()
    else:
        do_sync()
